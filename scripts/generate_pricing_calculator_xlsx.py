#!/usr/bin/env python3
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "deliverables" / "business-kit" / "editable-xlsx"
OUT_FILE = OUT_DIR / "rito-pricing-calculator.xlsx"

BRAND_DARK = "0D2430"
BRAND = "173847"
BRAND_MID = "315161"
BG = "F4F0E9"
SURFACE = "FFFDF9"
ACCENT = "B89163"
INK = "152733"
INK_SOFT = "5D6B72"
INPUT_FILL = "FFF2CC"
FORMULA_FILL = "F7F3EC"
WHITE = "FFFFFF"
LINE = "D7D1CA"

THIN = Side(style="thin", color=LINE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

RATE_ROWS = [
    ("Análise e descoberta", 150),
    ("Planejamento da solução", 150),
    ("Desenvolvimento", 180),
    ("Testes e ajustes", 180),
    ("Entrega e orientação", 150),
    ("Suporte inicial", 180),
    ("Treinamento adicional", 150),
]

COMPLEXITY_ROWS = [
    ("Baixa", 0.95),
    ("Média", 1.00),
    ("Alta", 1.15),
]

URGENCY_ROWS = [
    ("Normal", 1.00),
    ("Alta", 1.10),
    ("Crítica", 1.20),
]

INTEGRATION_ROWS = [
    ("Nenhuma", 1.00),
    ("Simples", 1.08),
    ("Múltiplas", 1.15),
]

DOCUMENTATION_ROWS = [
    ("Não", 1.00),
    ("Sim", 1.05),
]

SERVICE_ROWS = [
    ("Automação simples", 800, 2500),
    ("Automação com integração", 1800, 4500),
    ("Painel ou controle interno simples", 2500, 6000),
    ("Ferramenta operacional pequena", 4500, 9000),
    ("Sistema sob medida de pequeno porte", 8000, 18000),
    ("Projeto médio com integrações", 15000, 35000),
    ("Evolução ou manutenção técnica", "", ""),
]

PACKAGE_ROWS = [
    ("Essencial", 0.92, "Projetos enxutos, escopo mais direto e risco baixo."),
    ("Operacional", 1.00, "Base padrão da RITO para propostas equilibradas."),
    ("Sob Medida", 1.12, "Projetos com maior personalização, criticidade ou responsabilidade."),
]

SCENARIO_ROWS = [
    ("Mínimo defensável", 0.10, 1.00, "Protege margem mínima sem subprecificar."),
    ("Ideal comercial", 0.15, 1.00, "Referência principal para a maior parte das propostas."),
    ("Premium estratégico", 0.22, 1.06, "Aplicável quando há alto valor percebido, risco ou urgência relevante."),
]

INSTALLMENT_ROWS = [
    ("2 marcos", 0.50, 0.50, 0.00, 0.00),
    ("3 marcos", 0.40, 0.30, 0.30, 0.00),
    ("4 marcos", 0.30, 0.25, 0.25, 0.20),
]

GLOBAL_ROWS = [
    ("Margem padrão", 0.15),
    ("Entrada sugerida", 0.40),
    ("Parcela intermediária", 0.30),
    ("Saldo final", 0.30),
]


def style_cell(cell, *, fill=None, font=None, alignment=None, border=True, number_format=None):
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = BORDER
    if number_format:
        cell.number_format = number_format


def set_col_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def add_table_header(ws, cells):
    for coord, value in cells.items():
        ws[coord] = value
        style_cell(
            ws[coord],
            fill=BRAND,
            font=Font(name="Manrope", size=10, bold=True, color=WHITE),
            alignment=Alignment(horizontal="center", vertical="center"),
        )


def build_param_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Parametros")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:AG1")
    ws["A1"] = "Parâmetros da Calculadora de Precificação"
    style_cell(
        ws["A1"],
        fill=BRAND_DARK,
        font=Font(name="Sora", size=14, bold=True, color=WHITE),
        alignment=Alignment(horizontal="center", vertical="center"),
    )

    ws["A2"] = "Edite esta aba apenas quando as regras internas da RITO forem formalmente revisadas."
    style_cell(
        ws["A2"],
        font=Font(name="Manrope", size=10, color=INK_SOFT, italic=True),
        alignment=Alignment(horizontal="left"),
        border=False,
    )

    add_table_header(ws, {"A4": "Etapa", "B4": "Valor/hora base"})
    for index, (label, value) in enumerate(RATE_ROWS, start=5):
        ws[f"A{index}"] = label
        ws[f"B{index}"] = value
        style_cell(ws[f"A{index}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK))
        style_cell(ws[f"B{index}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK), number_format='R$ #,##0.00')

    add_table_header(ws, {"D4": "Complexidade", "E4": "Multiplicador"})
    for index, (label, value) in enumerate(COMPLEXITY_ROWS, start=5):
        ws[f"D{index}"] = label
        ws[f"E{index}"] = value
        style_cell(ws[f"D{index}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK))
        style_cell(ws[f"E{index}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK), number_format='0.00x')

    add_table_header(ws, {"G4": "Urgência", "H4": "Multiplicador"})
    for index, (label, value) in enumerate(URGENCY_ROWS, start=5):
        ws[f"G{index}"] = label
        ws[f"H{index}"] = value
        style_cell(ws[f"G{index}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK))
        style_cell(ws[f"H{index}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK), number_format='0.00x')

    add_table_header(ws, {"J4": "Integrações", "K4": "Multiplicador"})
    for index, (label, value) in enumerate(INTEGRATION_ROWS, start=5):
        ws[f"J{index}"] = label
        ws[f"K{index}"] = value
        style_cell(ws[f"J{index}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK))
        style_cell(ws[f"K{index}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK), number_format='0.00x')

    add_table_header(ws, {"M4": "Documentação extra", "N4": "Multiplicador"})
    for index, (label, value) in enumerate(DOCUMENTATION_ROWS, start=5):
        ws[f"M{index}"] = label
        ws[f"N{index}"] = value
        style_cell(ws[f"M{index}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK))
        style_cell(ws[f"N{index}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK), number_format='0.00x')

    add_table_header(ws, {"P4": "Tipo de serviço", "Q4": "Faixa mínima", "R4": "Faixa máxima"})
    for index, (service, min_value, max_value) in enumerate(SERVICE_ROWS, start=5):
        ws[f"P{index}"] = service
        ws[f"Q{index}"] = min_value
        ws[f"R{index}"] = max_value
        style_cell(ws[f"P{index}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK))
        style_cell(ws[f"Q{index}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK), number_format='R$ #,##0.00')
        style_cell(ws[f"R{index}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK), number_format='R$ #,##0.00')

    add_table_header(ws, {"T4": "Pacote", "U4": "Multiplicador", "V4": "Descrição"})
    for index, (label, multiplier, description) in enumerate(PACKAGE_ROWS, start=5):
        ws[f"T{index}"] = label
        ws[f"U{index}"] = multiplier
        ws[f"V{index}"] = description
        style_cell(ws[f"T{index}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK))
        style_cell(ws[f"U{index}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK), number_format='0.00x')
        style_cell(ws[f"V{index}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK), alignment=Alignment(wrap_text=True))

    add_table_header(ws, {"X4": "Cenário", "Y4": "Margem", "Z4": "Multiplicador", "AA4": "Uso sugerido"})
    for index, (label, margin, multiplier, note) in enumerate(SCENARIO_ROWS, start=5):
        ws[f"X{index}"] = label
        ws[f"Y{index}"] = margin
        ws[f"Z{index}"] = multiplier
        ws[f"AA{index}"] = note
        style_cell(ws[f"X{index}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK))
        style_cell(ws[f"Y{index}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK), number_format='0.00%')
        style_cell(ws[f"Z{index}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK), number_format='0.00x')
        style_cell(ws[f"AA{index}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK), alignment=Alignment(wrap_text=True))

    add_table_header(ws, {"AC4": "Modelo", "AD4": "Parcela 1", "AE4": "Parcela 2", "AF4": "Parcela 3", "AG4": "Parcela 4"})
    for index, (label, p1, p2, p3, p4) in enumerate(INSTALLMENT_ROWS, start=5):
        ws[f"AC{index}"] = label
        ws[f"AD{index}"] = p1
        ws[f"AE{index}"] = p2
        ws[f"AF{index}"] = p3
        ws[f"AG{index}"] = p4
        style_cell(ws[f"AC{index}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK))
        for col in ("AD", "AE", "AF", "AG"):
            style_cell(ws[f"{col}{index}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK), number_format='0.00%')

    add_table_header(ws, {"J11": "Parâmetro global", "K11": "Valor"})
    for index, (label, value) in enumerate(GLOBAL_ROWS, start=12):
        ws[f"J{index}"] = label
        ws[f"K{index}"] = value
        style_cell(ws[f"J{index}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK))
        style_cell(ws[f"K{index}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK), number_format='0.00%')

    set_col_widths(
        ws,
        {
            1: 28,
            2: 18,
            4: 16,
            5: 14,
            7: 16,
            8: 14,
            10: 18,
            11: 14,
            13: 20,
            14: 14,
            16: 34,
            17: 16,
            18: 16,
            20: 16,
            21: 14,
            22: 46,
            24: 18,
            25: 12,
            26: 12,
            27: 46,
            29: 14,
            30: 12,
            31: 12,
            32: 12,
            33: 12,
        },
    )


def add_main_input_row(ws, row, label, default="", number_format=None):
    ws[f"A{row}"] = label
    style_cell(ws[f"A{row}"], fill=BRAND, font=Font(name="Manrope", size=10, bold=True, color=WHITE))
    ws[f"B{row}"] = default
    style_cell(
        ws[f"B{row}"],
        fill=INPUT_FILL,
        font=Font(name="Manrope", size=10, color=INK),
        alignment=Alignment(horizontal="left"),
        number_format=number_format,
    )


def build_calculator_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Calculadora"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A18"

    ws.merge_cells("A1:F1")
    ws["A1"] = "Calculadora de Precificação | RITO Sistemas"
    style_cell(
        ws["A1"],
        fill=BRAND_DARK,
        font=Font(name="Sora", size=16, bold=True, color=WHITE),
        alignment=Alignment(horizontal="center", vertical="center"),
    )

    ws.merge_cells("A2:F2")
    ws["A2"] = "Preencha principalmente as células amarelas. A planilha calcula pacote, cenários de preço e parcelamentos sugeridos automaticamente."
    style_cell(
        ws["A2"],
        fill=BG,
        font=Font(name="Manrope", size=10, color=INK),
        alignment=Alignment(horizontal="center", vertical="center"),
    )

    add_main_input_row(ws, 4, "Cliente")
    add_main_input_row(ws, 5, "Projeto")
    add_main_input_row(ws, 6, "Tipo de serviço", "Sistema sob medida de pequeno porte")
    add_main_input_row(ws, 7, "Pacote comercial", "Operacional")
    add_main_input_row(ws, 8, "Complexidade", "Média")
    add_main_input_row(ws, 9, "Urgência", "Normal")
    add_main_input_row(ws, 10, "Integrações", "Nenhuma")
    add_main_input_row(ws, 11, "Documentação extra", "Não")
    add_main_input_row(ws, 12, "Desconto (%)", 0, '0.00%')
    add_main_input_row(ws, 13, "Custos de terceiros", 0, 'R$ #,##0.00')
    add_main_input_row(ws, 14, "Cenário preferencial", "Ideal comercial")

    right_labels = {
        "D4": "Faixa mínima",
        "D5": "Faixa máxima",
        "D6": "Status do ideal",
        "D7": "Multiplicador do pacote",
        "D8": "Descrição do pacote",
        "D9": "Leitura do cenário",
        "D10": "Parcelamento sugerido",
    }
    for cell, label in right_labels.items():
        ws[cell] = label
        style_cell(ws[cell], fill=BRAND_MID, font=Font(name="Manrope", size=10, bold=True, color=WHITE))

    ws["E4"] = '=IFERROR(INDEX(Parametros!$Q$5:$Q$11,MATCH($B$6,Parametros!$P$5:$P$11,0)),"")'
    ws["E5"] = '=IFERROR(INDEX(Parametros!$R$5:$R$11,MATCH($B$6,Parametros!$P$5:$P$11,0)),"")'
    ws["E6"] = '=IF(OR(E4="",E5=""),"Faixa aberta / usar hora técnica",IF($B$36<E4,"Abaixo da faixa de referência",IF($B$36>E5,"Acima da faixa de referência","Dentro da faixa de referência")))'
    ws["E7"] = '=IFERROR(INDEX(Parametros!$U$5:$U$7,MATCH($B$7,Parametros!$T$5:$T$7,0)),1)'
    ws["E8"] = '=IFERROR(INDEX(Parametros!$V$5:$V$7,MATCH($B$7,Parametros!$T$5:$T$7,0)),"")'
    ws["E9"] = '=IFERROR(INDEX(Parametros!$AA$5:$AA$7,MATCH($B$14,Parametros!$X$5:$X$7,0)),"")'
    ws["E10"] = '=Parcelamento!$B$6'
    for cell in ("E4", "E5", "E6", "E7", "E8", "E9", "E10"):
        style_cell(
            ws[cell],
            fill=FORMULA_FILL,
            font=Font(name="Manrope", size=10, color=INK),
            alignment=Alignment(wrap_text=True),
            number_format='R$ #,##0.00' if cell in ("E4", "E5") else '0.00x' if cell == "E7" else None,
        )

    headers = ["Etapa", "Horas", "Valor/hora base", "Subtotal", "Observação"]
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(17, col_index, header)
        style_cell(
            cell,
            fill=BRAND_DARK,
            font=Font(name="Manrope", size=10, bold=True, color=WHITE),
            alignment=Alignment(horizontal="center"),
        )

    for row_index, (label, _) in enumerate(RATE_ROWS, start=18):
        ws[f"A{row_index}"] = label
        ws[f"B{row_index}"] = 0
        ws[f"C{row_index}"] = f'=IFERROR(INDEX(Parametros!$B$5:$B$11,MATCH(A{row_index},Parametros!$A$5:$A$11,0)),0)'
        ws[f"D{row_index}"] = f'=B{row_index}*C{row_index}'
        ws[f"E{row_index}"] = ""

        style_cell(ws[f"A{row_index}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK))
        style_cell(ws[f"B{row_index}"], fill=INPUT_FILL, font=Font(name="Manrope", size=10, color=INK), number_format='0.00')
        style_cell(ws[f"C{row_index}"], fill=FORMULA_FILL, font=Font(name="Manrope", size=10, color=INK), number_format='R$ #,##0.00')
        style_cell(ws[f"D{row_index}"], fill=FORMULA_FILL, font=Font(name="Manrope", size=10, color=INK), number_format='R$ #,##0.00')
        style_cell(ws[f"E{row_index}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK))

    summary_rows = [
        (27, "Horas totais", '=SUM(B18:B24)', '0.00'),
        (28, "Subtotal base", '=SUM(D18:D24)', 'R$ #,##0.00'),
        (29, "Multiplicador de complexidade", '=IFERROR(INDEX(Parametros!$E$5:$E$7,MATCH($B$8,Parametros!$D$5:$D$7,0)),1)', '0.00x'),
        (30, "Multiplicador de urgência", '=IFERROR(INDEX(Parametros!$H$5:$H$7,MATCH($B$9,Parametros!$G$5:$G$7,0)),1)', '0.00x'),
        (31, "Multiplicador de integrações", '=IFERROR(INDEX(Parametros!$K$5:$K$7,MATCH($B$10,Parametros!$J$5:$J$7,0)),1)', '0.00x'),
        (32, "Multiplicador de documentação", '=IFERROR(INDEX(Parametros!$N$5:$N$6,MATCH($B$11,Parametros!$M$5:$M$6,0)),1)', '0.00x'),
        (33, "Subtotal técnico ajustado", '=B28*B29*B30*B31*B32', 'R$ #,##0.00'),
        (34, "Base comercial do pacote", '=B33*IFERROR(INDEX(Parametros!$U$5:$U$7,MATCH($B$7,Parametros!$T$5:$T$7,0)),1)', 'R$ #,##0.00'),
        (35, "Mínimo defensável", '=Cenarios!H8', 'R$ #,##0.00'),
        (36, "Ideal comercial", '=Cenarios!H9', 'R$ #,##0.00'),
        (37, "Premium estratégico", '=Cenarios!H10', 'R$ #,##0.00'),
        (38, "Preço do cenário preferencial", '=IFERROR(INDEX(Cenarios!$H$8:$H$10,MATCH($B$14,Cenarios!$A$8:$A$10,0)),0)', 'R$ #,##0.00'),
        (39, "Valor/hora ideal", '=Cenarios!I9', 'R$ #,##0.00'),
        (40, "Modelo de parcelamento sugerido", '=Parcelamento!$B$6', None),
    ]

    for row_index, label, formula, num_format in summary_rows:
        fill = BRAND if row_index >= 35 else BRAND_MID
        ws[f"A{row_index}"] = label
        ws[f"B{row_index}"] = formula
        style_cell(ws[f"A{row_index}"], fill=fill, font=Font(name="Manrope", size=10, bold=True, color=WHITE))
        style_cell(
            ws[f"B{row_index}"],
            fill=FORMULA_FILL,
            font=Font(name="Manrope", size=10, color=INK),
            number_format=num_format,
        )

    note_rows = {
        "D27": "Como ler",
        "E27": "Use as horas por etapa como base. O pacote comercial ajusta a proposta ao posicionamento da oferta.",
        "D28": "Pacotes",
        "E28": "Essencial reduz a proposta para algo mais enxuto; Operacional é a base padrão; Sob Medida protege escopo de maior personalização.",
        "D29": "Cenários",
        "E29": "Mínimo = limite defensável, Ideal = referência principal, Premium = quando valor percebido, risco ou urgência forem maiores.",
        "D30": "Desconto",
        "E30": "Só aplicar desconto com justificativa comercial clara. Sempre preferir revisar escopo antes de baixar preço.",
        "D31": "Faixa de serviço",
        "E31": "Se o ideal ficar muito fora da faixa de referência, revisar escopo, pacote, risco ou tipo de serviço selecionado.",
    }
    for cell, value in note_rows.items():
        fill = BRAND_MID if cell.startswith("D") else SURFACE
        font = Font(name="Manrope", size=10, bold=True, color=WHITE) if cell.startswith("D") else Font(name="Manrope", size=10, color=INK)
        ws[cell] = value
        style_cell(ws[cell], fill=fill, font=font, alignment=Alignment(wrap_text=True))

    set_col_widths(ws, {1: 34, 2: 18, 3: 16, 4: 18, 5: 34, 6: 3})

    validations = [
        (DataValidation(type="list", formula1="=Parametros!$P$5:$P$11", allow_blank=False), "B6"),
        (DataValidation(type="list", formula1="=Parametros!$T$5:$T$7", allow_blank=False), "B7"),
        (DataValidation(type="list", formula1="=Parametros!$D$5:$D$7", allow_blank=False), "B8"),
        (DataValidation(type="list", formula1="=Parametros!$G$5:$G$7", allow_blank=False), "B9"),
        (DataValidation(type="list", formula1="=Parametros!$J$5:$J$7", allow_blank=False), "B10"),
        (DataValidation(type="list", formula1="=Parametros!$M$5:$M$6", allow_blank=False), "B11"),
        (DataValidation(type="list", formula1="=Parametros!$X$5:$X$7", allow_blank=False), "B14"),
    ]
    for validation, cell in validations:
        validation.error = "Selecione um valor da lista."
        validation.prompt = "Escolha uma opção predefinida."
        ws.add_data_validation(validation)
        validation.add(ws[cell])


def build_scenarios_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Cenarios")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:K1")
    ws["A1"] = "Comparação de Cenários e Pacotes"
    style_cell(
        ws["A1"],
        fill=BRAND_DARK,
        font=Font(name="Sora", size=15, bold=True, color=WHITE),
        alignment=Alignment(horizontal="center"),
    )

    ws["A2"] = "Esta aba compara os cenários de preço para o pacote selecionado e mostra a diferença entre Essencial, Operacional e Sob Medida no cenário ideal."
    style_cell(ws["A2"], fill=BG, font=Font(name="Manrope", size=10, color=INK), alignment=Alignment(wrap_text=True))

    summary_labels = {
        "A4": "Projeto",
        "A5": "Tipo de serviço",
        "D4": "Pacote selecionado",
        "D5": "Base comercial do pacote",
    }
    for cell, value in summary_labels.items():
        ws[cell] = value
        style_cell(ws[cell], fill=BRAND_MID, font=Font(name="Manrope", size=10, bold=True, color=WHITE))
    ws["B4"] = "=Calculadora!B5"
    ws["B5"] = "=Calculadora!B6"
    ws["E4"] = "=Calculadora!B7"
    ws["E5"] = "=Calculadora!B34"
    style_cell(ws["B4"], fill=FORMULA_FILL, font=Font(name="Manrope", size=10, color=INK))
    style_cell(ws["B5"], fill=FORMULA_FILL, font=Font(name="Manrope", size=10, color=INK))
    style_cell(ws["E4"], fill=FORMULA_FILL, font=Font(name="Manrope", size=10, color=INK))
    style_cell(ws["E5"], fill=FORMULA_FILL, font=Font(name="Manrope", size=10, color=INK), number_format='R$ #,##0.00')

    headers = {
        "A7": "Cenário",
        "B7": "Margem",
        "C7": "Mult. comercial",
        "D7": "Base do pacote",
        "E7": "Serviços antes do desconto",
        "F7": "Desconto",
        "G7": "Custos terceiros",
        "H7": "Total final",
        "I7": "Valor/hora",
        "J7": "Status vs faixa",
        "K7": "Uso sugerido",
    }
    add_table_header(ws, headers)

    for row_index in range(8, 11):
        param_row = row_index - 3
        ws[f"A{row_index}"] = f"=Parametros!X{param_row}"
        ws[f"B{row_index}"] = f"=Parametros!Y{param_row}"
        ws[f"C{row_index}"] = f"=Parametros!Z{param_row}"
        ws[f"D{row_index}"] = "=Calculadora!$B$34"
        ws[f"E{row_index}"] = f"=(D{row_index}*C{row_index})*(1+B{row_index})"
        ws[f"F{row_index}"] = f"=E{row_index}*Calculadora!$B$12"
        ws[f"G{row_index}"] = "=Calculadora!$B$13"
        ws[f"H{row_index}"] = f"=E{row_index}-F{row_index}+G{row_index}"
        ws[f"I{row_index}"] = f'=IFERROR(H{row_index}/Calculadora!$B$27,0)'
        ws[f"J{row_index}"] = (
            f'=IF(OR(Calculadora!$E$4="",Calculadora!$E$5=""),"Faixa aberta / usar hora técnica",'
            f'IF(H{row_index}<Calculadora!$E$4,"Abaixo da faixa",IF(H{row_index}>Calculadora!$E$5,"Acima da faixa","Dentro da faixa")))'
        )
        ws[f"K{row_index}"] = f"=Parametros!AA{param_row}"

        for col in ("A", "J", "K"):
            style_cell(ws[f"{col}{row_index}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK), alignment=Alignment(wrap_text=True))
        for col in ("B", "C"):
            style_cell(ws[f"{col}{row_index}"], fill=FORMULA_FILL, font=Font(name="Manrope", size=10, color=INK), number_format='0.00x' if col == "C" else '0.00%')
        for col in ("D", "E", "F", "G", "H", "I"):
            style_cell(ws[f"{col}{row_index}"], fill=FORMULA_FILL, font=Font(name="Manrope", size=10, color=INK), number_format='R$ #,##0.00')

    add_table_header(ws, {"A13": "Pacote", "B13": "Multiplicador", "C13": "Total ideal", "D13": "Leitura"})
    for row_index in range(14, 17):
        param_row = row_index - 9
        ws[f"A{row_index}"] = f"=Parametros!T{param_row}"
        ws[f"B{row_index}"] = f"=Parametros!U{param_row}"
        ws[f"C{row_index}"] = (
            f'=(((Calculadora!$B$33*B{row_index})*INDEX(Parametros!$Z$5:$Z$7,MATCH("Ideal comercial",Parametros!$X$5:$X$7,0)))'
            f'*(1+INDEX(Parametros!$Y$5:$Y$7,MATCH("Ideal comercial",Parametros!$X$5:$X$7,0))))'
            f'-((((Calculadora!$B$33*B{row_index})*INDEX(Parametros!$Z$5:$Z$7,MATCH("Ideal comercial",Parametros!$X$5:$X$7,0)))'
            f'*(1+INDEX(Parametros!$Y$5:$Y$7,MATCH("Ideal comercial",Parametros!$X$5:$X$7,0))))*Calculadora!$B$12)'
            f'+Calculadora!$B$13'
        )
        ws[f"D{row_index}"] = f"=Parametros!V{param_row}"
        style_cell(ws[f"A{row_index}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK))
        style_cell(ws[f"B{row_index}"], fill=FORMULA_FILL, font=Font(name="Manrope", size=10, color=INK), number_format='0.00x')
        style_cell(ws[f"C{row_index}"], fill=FORMULA_FILL, font=Font(name="Manrope", size=10, color=INK), number_format='R$ #,##0.00')
        style_cell(ws[f"D{row_index}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK), alignment=Alignment(wrap_text=True))

    set_col_widths(ws, {1: 22, 2: 12, 3: 13, 4: 17, 5: 20, 6: 14, 7: 16, 8: 16, 9: 14, 10: 20, 11: 42})


def build_installment_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Parcelamento")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:K1")
    ws["A1"] = "Simulação de Parcelamento por Marcos"
    style_cell(
        ws["A1"],
        fill=BRAND_DARK,
        font=Font(name="Sora", size=15, bold=True, color=WHITE),
        alignment=Alignment(horizontal="center"),
    )

    ws["A2"] = "Escolha o cenário-base e compare os modelos de 2, 3 e 4 marcos. O total vem automaticamente da aba de cenários."
    style_cell(ws["A2"], fill=BG, font=Font(name="Manrope", size=10, color=INK), alignment=Alignment(wrap_text=True))

    top_labels = {
        "A4": "Cenário-base",
        "A5": "Total da proposta",
        "A6": "Modelo sugerido",
        "A7": "Valor/hora do cenário",
        "A8": "Status vs faixa",
    }
    for cell, value in top_labels.items():
        ws[cell] = value
        style_cell(ws[cell], fill=BRAND, font=Font(name="Manrope", size=10, bold=True, color=WHITE))

    ws["B4"] = '=Calculadora!$B$14'
    ws["B5"] = '=IFERROR(INDEX(Cenarios!$H$8:$H$10,MATCH($B$4,Cenarios!$A$8:$A$10,0)),0)'
    ws["B6"] = '=IF(B5<=5000,"2 marcos",IF(B5<=15000,"3 marcos","4 marcos"))'
    ws["B7"] = '=IFERROR(INDEX(Cenarios!$I$8:$I$10,MATCH($B$4,Cenarios!$A$8:$A$10,0)),0)'
    ws["B8"] = '=IFERROR(INDEX(Cenarios!$J$8:$J$10,MATCH($B$4,Cenarios!$A$8:$A$10,0)),"")'

    style_cell(ws["B4"], fill=FORMULA_FILL, font=Font(name="Manrope", size=10, color=INK))
    style_cell(ws["B5"], fill=FORMULA_FILL, font=Font(name="Manrope", size=10, color=INK), number_format='R$ #,##0.00')
    style_cell(ws["B6"], fill=FORMULA_FILL, font=Font(name="Manrope", size=10, color=INK))
    style_cell(ws["B7"], fill=FORMULA_FILL, font=Font(name="Manrope", size=10, color=INK), number_format='R$ #,##0.00')
    style_cell(ws["B8"], fill=FORMULA_FILL, font=Font(name="Manrope", size=10, color=INK), alignment=Alignment(wrap_text=True))

    sections = [
        ("A10", "Modelo 2 marcos", 5, ["Entrada", "Entrega final"]),
        ("E10", "Modelo 3 marcos", 6, ["Entrada", "Marco intermediário", "Entrega final"]),
        ("I10", "Modelo 4 marcos", 7, ["Entrada", "Planejamento aprovado", "Primeira entrega", "Entrega final"]),
    ]

    for top_left, title, param_row, labels in sections:
        col = top_left[0]
        start_row = int(top_left[1:])
        value_col = chr(ord(col) + 1)
        amount_col = chr(ord(col) + 2)

        add_table_header(ws, {f"{col}{start_row}": title, f"{value_col}{start_row}": "%", f"{amount_col}{start_row}": "Valor"})

        for offset, label in enumerate(labels, start=1):
            row = start_row + offset
            percent_formula_col = ["AD", "AE", "AF", "AG"][offset - 1]
            ws[f"{col}{row}"] = label
            ws[f"{value_col}{row}"] = f"=Parametros!${percent_formula_col}${param_row}"
            ws[f"{amount_col}{row}"] = f"=${value_col}{row}*$B$5"

            style_cell(ws[f"{col}{row}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK))
            style_cell(ws[f"{value_col}{row}"], fill=FORMULA_FILL, font=Font(name="Manrope", size=10, color=INK), number_format='0.00%')
            style_cell(ws[f"{amount_col}{row}"], fill=FORMULA_FILL, font=Font(name="Manrope", size=10, color=INK), number_format='R$ #,##0.00')

        total_row = start_row + len(labels) + 1
        ws[f"{col}{total_row}"] = "Total"
        ws[f"{value_col}{total_row}"] = f"=SUM({value_col}{start_row + 1}:{value_col}{start_row + len(labels)})"
        ws[f"{amount_col}{total_row}"] = f"=SUM({amount_col}{start_row + 1}:{amount_col}{start_row + len(labels)})"
        style_cell(ws[f"{col}{total_row}"], fill=BRAND_MID, font=Font(name="Manrope", size=10, bold=True, color=WHITE))
        style_cell(ws[f"{value_col}{total_row}"], fill=FORMULA_FILL, font=Font(name="Manrope", size=10, color=INK), number_format='0.00%')
        style_cell(ws[f"{amount_col}{total_row}"], fill=FORMULA_FILL, font=Font(name="Manrope", size=10, color=INK), number_format='R$ #,##0.00')

    ws["A18"] = "Leitura"
    ws["B18"] = "2 marcos funciona melhor para propostas menores. 3 marcos é a estrutura padrão mais equilibrada. 4 marcos ajuda a proteger caixa e risco em projetos maiores."
    style_cell(ws["A18"], fill=BRAND_MID, font=Font(name="Manrope", size=10, bold=True, color=WHITE))
    style_cell(ws["B18"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK), alignment=Alignment(wrap_text=True))

    set_col_widths(ws, {1: 24, 2: 12, 3: 16, 5: 24, 6: 12, 7: 16, 9: 24, 10: 12, 11: 16})


def build_summary_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Resumo-Proposta")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A12"
    ws.print_area = "A1:F36"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    ws.merge_cells("A1:F1")
    ws["A1"] = "Resumo de Orçamento e Proposta"
    style_cell(
        ws["A1"],
        fill=BRAND_DARK,
        font=Font(name="Sora", size=16, bold=True, color=WHITE),
        alignment=Alignment(horizontal="center", vertical="center"),
    )

    ws.merge_cells("A2:F2")
    ws["A2"] = "Esta aba foi desenhada para revisão rápida e exportação em PDF direto pelo Excel."
    style_cell(ws["A2"], fill=BG, font=Font(name="Manrope", size=10, color=INK), alignment=Alignment(horizontal="center"))

    info_pairs = [
        ("A4", "Cliente", "B4", "=Calculadora!B4", None),
        ("A5", "Projeto", "B5", "=Calculadora!B5", None),
        ("A6", "Tipo de serviço", "B6", "=Calculadora!B6", None),
        ("A7", "Pacote comercial", "B7", "=Calculadora!B7", None),
        ("A8", "Cenário adotado", "B8", "=Calculadora!B14", None),
        ("A9", "Investimento sugerido", "B9", "=Calculadora!B38", 'R$ #,##0.00'),
        ("A10", "Horas estimadas", "B10", "=Calculadora!B27", '0.00'),
        ("D4", "Prazo estimado", "E4", "", None),
        ("D5", "Validade da proposta", "E5", "", None),
        ("D6", "Data de emissão", "E6", "=TODAY()", 'dd/mm/yyyy'),
        ("D7", "Modelo de parcelamento", "E7", "=Parcelamento!B6", None),
        ("D8", "Parcela 1", "E8", '=IF(E7="2 marcos",Parcelamento!C11,IF(E7="3 marcos",Parcelamento!G11,Parcelamento!K11))', 'R$ #,##0.00'),
        ("D9", "Parcela 2", "E9", '=IF(E7="2 marcos",Parcelamento!C12,IF(E7="3 marcos",Parcelamento!G12,Parcelamento!K12))', 'R$ #,##0.00'),
        ("D10", "Parcela 3/4", "E10", '=IF(E7="2 marcos","",IF(E7="3 marcos",Parcelamento!G13,Parcelamento!K13&" / "&TEXT(Parcelamento!K14,"R$ #,##0.00")))', None),
    ]

    for label_cell, label, value_cell, formula_or_value, num_format in info_pairs:
        ws[label_cell] = label
        style_cell(ws[label_cell], fill=BRAND_MID, font=Font(name="Manrope", size=10, bold=True, color=WHITE))
        ws[value_cell] = formula_or_value
        fill = INPUT_FILL if formula_or_value == "" else FORMULA_FILL
        style_cell(ws[value_cell], fill=fill, font=Font(name="Manrope", size=10, color=INK), alignment=Alignment(wrap_text=True), number_format=num_format)

    section_headers = {
        "A12": "Resumo da demanda",
        "A17": "Solução sugerida",
        "A22": "Composição do investimento",
        "D22": "Observações comerciais",
        "D29": "Próximo passo sugerido",
    }
    for cell, value in section_headers.items():
        ws[cell] = value
        style_cell(ws[cell], fill=BRAND, font=Font(name="Manrope", size=10, bold=True, color=WHITE))

    for merge_range in ("A13:F15", "A18:F20", "D23:F27", "D30:F33"):
        ws.merge_cells(merge_range)
        top_left = ws[merge_range.split(":")[0]]
        style_cell(top_left, fill=INPUT_FILL, font=Font(name="Manrope", size=10, color=INK), alignment=Alignment(wrap_text=True, vertical="top"))

    ws["A13"] = ""
    ws["A18"] = ""
    ws["D23"] = "Use este espaço para registrar premissas, limites de escopo, dependências ou ressalvas importantes antes de exportar."
    ws["D30"] = "Após validar este resumo, leve o cenário escolhido para a proposta comercial final."

    add_table_header(ws, {"A23": "Etapa", "B23": "Horas", "C23": "Subtotal"})
    for row_offset, calc_row in enumerate(range(18, 25), start=24):
        ws[f"A{row_offset}"] = f"=Calculadora!A{calc_row}"
        ws[f"B{row_offset}"] = f"=Calculadora!B{calc_row}"
        ws[f"C{row_offset}"] = f"=Calculadora!D{calc_row}"
        style_cell(ws[f"A{row_offset}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK))
        style_cell(ws[f"B{row_offset}"], fill=FORMULA_FILL, font=Font(name="Manrope", size=10, color=INK), number_format='0.00')
        style_cell(ws[f"C{row_offset}"], fill=FORMULA_FILL, font=Font(name="Manrope", size=10, color=INK), number_format='R$ #,##0.00')

    ws["A31"] = "Total do cenário"
    ws["B31"] = "=Calculadora!B38"
    style_cell(ws["A31"], fill=BRAND_DARK, font=Font(name="Manrope", size=11, bold=True, color=WHITE))
    style_cell(ws["B31"], fill=FORMULA_FILL, font=Font(name="Manrope", size=11, bold=True, color=INK), number_format='R$ #,##0.00')

    ws["A35"] = "Exportação"
    ws["B35"] = "Depois de revisar esta aba, exporte somente `Resumo-Proposta` em PDF pelo Excel para gerar um orçamento-resumo limpo."
    style_cell(ws["A35"], fill=BRAND_MID, font=Font(name="Manrope", size=10, bold=True, color=WHITE))
    ws.merge_cells("B35:F35")
    style_cell(ws["B35"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK), alignment=Alignment(wrap_text=True))

    set_col_widths(ws, {1: 22, 2: 20, 3: 16, 4: 22, 5: 24, 6: 18})


def build_history_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Historico")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A7"

    ws.merge_cells("A1:J1")
    ws["A1"] = "Histórico de Estimado, Vendido e Realizado"
    style_cell(
        ws["A1"],
        fill=BRAND_DARK,
        font=Font(name="Sora", size=15, bold=True, color=WHITE),
        alignment=Alignment(horizontal="center"),
    )

    ws["A2"] = "Use esta aba após fechar a proposta e ao longo da execução do projeto para comparar horas planejadas, vendidas e reais."
    style_cell(ws["A2"], fill=BG, font=Font(name="Manrope", size=10, color=INK), alignment=Alignment(wrap_text=True))

    top_info = {
        "A4": "Projeto",
        "B4": "=Calculadora!B5",
        "D4": "Cenário vendido",
        "E4": "=Calculadora!B14",
        "G4": "Preço vendido",
        "H4": "=Calculadora!B38",
    }
    for cell, value in top_info.items():
        ws[cell] = value
        if cell in ("A4", "D4", "G4"):
            style_cell(ws[cell], fill=BRAND_MID, font=Font(name="Manrope", size=10, bold=True, color=WHITE))
        else:
            style_cell(ws[cell], fill=FORMULA_FILL, font=Font(name="Manrope", size=10, color=INK), number_format='R$ #,##0.00' if cell == "H4" else None)

    add_table_header(
        ws,
        {
            "A6": "Etapa",
            "B6": "Horas estimadas",
            "C6": "Horas vendidas",
            "D6": "Horas reais",
            "E6": "Desvio vendido x estimado",
            "F6": "Desvio real x vendido",
            "G6": "Valor/h base",
            "H6": "Valor vendido da etapa",
            "I6": "Custo real pela etapa",
            "J6": "Saldo por etapa",
        },
    )

    for row_offset, calc_row in enumerate(range(18, 25), start=7):
        ws[f"A{row_offset}"] = f"=Calculadora!A{calc_row}"
        ws[f"B{row_offset}"] = f"=Calculadora!B{calc_row}"
        ws[f"C{row_offset}"] = f"=B{row_offset}"
        ws[f"D{row_offset}"] = 0
        ws[f"E{row_offset}"] = f"=C{row_offset}-B{row_offset}"
        ws[f"F{row_offset}"] = f"=D{row_offset}-C{row_offset}"
        ws[f"G{row_offset}"] = f"=Calculadora!C{calc_row}"
        ws[f"H{row_offset}"] = f"=C{row_offset}*G{row_offset}"
        ws[f"I{row_offset}"] = f"=D{row_offset}*G{row_offset}"
        ws[f"J{row_offset}"] = f"=H{row_offset}-I{row_offset}"

        style_cell(ws[f"A{row_offset}"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK))
        style_cell(ws[f"B{row_offset}"], fill=FORMULA_FILL, font=Font(name="Manrope", size=10, color=INK), number_format='0.00')
        style_cell(ws[f"C{row_offset}"], fill=INPUT_FILL, font=Font(name="Manrope", size=10, color=INK), number_format='0.00')
        style_cell(ws[f"D{row_offset}"], fill=INPUT_FILL, font=Font(name="Manrope", size=10, color=INK), number_format='0.00')
        for col in ("E", "F"):
            style_cell(ws[f"{col}{row_offset}"], fill=FORMULA_FILL, font=Font(name="Manrope", size=10, color=INK), number_format='0.00')
        style_cell(ws[f"G{row_offset}"], fill=FORMULA_FILL, font=Font(name="Manrope", size=10, color=INK), number_format='R$ #,##0.00')
        style_cell(ws[f"H{row_offset}"], fill=FORMULA_FILL, font=Font(name="Manrope", size=10, color=INK), number_format='R$ #,##0.00')
        style_cell(ws[f"I{row_offset}"], fill=FORMULA_FILL, font=Font(name="Manrope", size=10, color=INK), number_format='R$ #,##0.00')
        style_cell(ws[f"J{row_offset}"], fill=FORMULA_FILL, font=Font(name="Manrope", size=10, color=INK), number_format='R$ #,##0.00')

    totals = [
        ("A16", "Totais"),
        ("B16", "=SUM(B7:B13)"),
        ("C16", "=SUM(C7:C13)"),
        ("D16", "=SUM(D7:D13)"),
        ("H16", "=SUM(H7:H13)"),
        ("I16", "=SUM(I7:I13)"),
        ("J16", "=SUM(J7:J13)"),
    ]
    for cell, value in totals:
        ws[cell] = value
        if cell == "A16":
            style_cell(ws[cell], fill=BRAND_DARK, font=Font(name="Manrope", size=10, bold=True, color=WHITE))
        else:
            style_cell(ws[cell], fill=FORMULA_FILL, font=Font(name="Manrope", size=10, bold=True, color=INK), number_format='0.00' if cell in ("B16", "C16", "D16") else 'R$ #,##0.00')

    ws["A18"] = "Leitura rápida"
    ws["B18"] = "Use `Horas vendidas` para refletir o que foi de fato combinado com o cliente. Use `Horas reais` ao longo da execução. O saldo por etapa ajuda a identificar onde o projeto ganhou ou perdeu margem operacional."
    style_cell(ws["A18"], fill=BRAND_MID, font=Font(name="Manrope", size=10, bold=True, color=WHITE))
    ws.merge_cells("B18:J18")
    style_cell(ws["B18"], fill=SURFACE, font=Font(name="Manrope", size=10, color=INK), alignment=Alignment(wrap_text=True))

    set_col_widths(ws, {1: 28, 2: 14, 3: 14, 4: 14, 5: 18, 6: 18, 7: 14, 8: 18, 9: 18, 10: 18})


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)

    build_calculator_sheet(wb)
    build_scenarios_sheet(wb)
    build_installment_sheet(wb)
    build_summary_sheet(wb)
    build_history_sheet(wb)
    build_param_sheet(wb)

    wb.save(OUT_FILE)
    print(OUT_FILE.relative_to(ROOT))


if __name__ == "__main__":
    main()
