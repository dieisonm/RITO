from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.properties import CalcProperties
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
BATCH_DIR = (
    ROOT
    / "operations"
    / "ai-os"
    / "growth"
    / "prospecting"
    / "territories"
    / "novo-hamburgo-rs"
    / "batches"
)
CSV_FILE = BATCH_DIR / "2026-04-21-batch-001-pilot.csv"
JSON_FILE = BATCH_DIR / "2026-04-21-batch-001-outreach-data.json"
OUT_FILE = BATCH_DIR / "2026-04-21-batch-001-outbound-tracker.xlsx"


BRAND = "173847"
BRAND_STRONG = "0D2430"
ACCENT = "B89163"
BG = "F4F0E9"
SURFACE = "FFFDF9"
INK = "152733"
INK_SOFT = "5D6B72"
LINE = "D9D9D9"
GREEN = "D9F2E3"
YELLOW = "FFF0CC"
RED = "F8D7DA"
BLUE = "DCEAF5"


def load_rows() -> list[dict[str, str]]:
    with CSV_FILE.open("r", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    with JSON_FILE.open("r", encoding="utf-8") as f:
        outreach = {item["company_id"]: item for item in json.load(f)}

    for row in rows:
        extra = outreach.get(row["company_id"], {})
        row["priority"] = extra.get("priority", "media")
        row["recommended_channel"] = extra.get(
            "recommended_channel", row.get("primary_contact_channel", "")
        )
        row["email_subject"] = extra.get("email_subject", "")
        row["email_body"] = extra.get("email_body", "")
        row["whatsapp_body"] = extra.get("whatsapp_body", "")
        row["outbound_status"] = "nao-iniciado"
        row["owner_human"] = ""
        row["first_touch_date"] = ""
        row["last_touch_date"] = ""
        row["next_follow_up_date"] = ""
        row["reply_status"] = ""
        row["notes"] = ""
    return rows


def style_cell(cell, *, fill=None, font=None, align=None, border=None):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = border


def autofit_like(ws):
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            widths[cell.column] = max(widths.get(cell.column, 0), min(length, 70))
    for col_idx, length in widths.items():
        ws.column_dimensions[chr(64 + col_idx)].width = max(12, min(length + 2, 40))


def add_table(ws, start: str, end: str, name: str):
    table = Table(displayName=name, ref=f"{start}:{end}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def build_dashboard(wb: Workbook, row_count: int):
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"

    title_fill = PatternFill("solid", fgColor=BRAND_STRONG)
    card_fill = PatternFill("solid", fgColor=SURFACE)
    header_fill = PatternFill("solid", fgColor=BRAND)
    border = Border(
        left=Side(style="thin", color=LINE),
        right=Side(style="thin", color=LINE),
        top=Side(style="thin", color=LINE),
        bottom=Side(style="thin", color=LINE),
    )

    ws.merge_cells("A1:H2")
    ws["A1"] = "RITO Sistemas - Outbound Tracker"
    style_cell(
        ws["A1"],
        fill=title_fill,
        font=Font(color="FFFFFF", bold=True, size=18),
        align=Alignment(horizontal="center", vertical="center"),
        border=border,
    )
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 10

    ws.merge_cells("A3:H3")
    ws["A3"] = (
        "Batch piloto de Novo Hamburgo/RS com base, drafts de outreach e acompanhamento operacional."
    )
    style_cell(
        ws["A3"],
        font=Font(color=INK_SOFT, italic=True),
        align=Alignment(horizontal="left"),
    )

    cards = [
        ("A5:B5", "A6:B7", "Total de empresas", f"=COUNTA(Leads!A2:A{row_count + 20})"),
        ("C5:D5", "C6:D7", "Prontas para contato", f'=COUNTIF(Leads!S2:S{row_count + 20},"ready")'),
        ("E5:F5", "E6:F7", "Status nao iniciado", f'=COUNTIF(Leads!X2:X{row_count + 20},"nao-iniciado")'),
        ("G5:H5", "G6:H7", "Leads com e-mail", f'=COUNTIF(Leads!F2:F{row_count + 20},"<>")'),
    ]

    for label_ref, value_ref, label, formula in cards:
        ws.merge_cells(label_ref)
        ws.merge_cells(value_ref)
        label_cell = ws[label_ref.split(":")[0]]
        value_cell = ws[value_ref.split(":")[0]]
        label_cell.value = label
        value_cell.value = formula
        style_cell(
            label_cell,
            fill=card_fill,
            font=Font(color=INK_SOFT, bold=True, size=10),
            align=Alignment(horizontal="center", vertical="center"),
            border=border,
        )
        style_cell(
            value_cell,
            fill=card_fill,
            font=Font(color=INK, bold=True, size=18),
            align=Alignment(horizontal="center", vertical="center"),
            border=border,
        )

    ws["A10"] = "Distribuicao por canal sugerido"
    style_cell(ws["A10"], fill=header_fill, font=Font(color="FFFFFF", bold=True))
    ws["A11"] = "Canal"
    ws["B11"] = "Qtd"
    style_cell(ws["A11"], fill=card_fill, font=Font(bold=True), border=border)
    style_cell(ws["B11"], fill=card_fill, font=Font(bold=True), border=border)

    channels = [
        "email",
        "whatsapp-manual",
        "formulario-ou-whatsapp-manual",
        "formulario",
    ]
    start_row = 12
    for idx, channel in enumerate(channels, start=start_row):
        ws[f"A{idx}"] = channel
        ws[f"B{idx}"] = f'=COUNTIF(Leads!O2:O{row_count + 20},"{channel}")'
        style_cell(ws[f"A{idx}"], border=border)
        style_cell(ws[f"B{idx}"], border=border)

    ws["D10"] = "Prioridade"
    ws["E10"] = "Qtd"
    style_cell(ws["D10"], fill=header_fill, font=Font(color="FFFFFF", bold=True))
    style_cell(ws["E10"], fill=header_fill, font=Font(color="FFFFFF", bold=True))
    priorities = ["alta", "media", "baixa"]
    for idx, priority in enumerate(priorities, start=11):
        ws[f"D{idx}"] = priority
        ws[f"E{idx}"] = f'=COUNTIF(Leads!N2:N{row_count + 20},"{priority}")'
        style_cell(ws[f"D{idx}"], border=border)
        style_cell(ws[f"E{idx}"], border=border)

    chart = BarChart()
    chart.title = "Leads por canal"
    chart.style = 10
    chart.y_axis.title = "Quantidade"
    chart.x_axis.title = "Canal"
    data = Reference(ws, min_col=2, min_row=11, max_row=15)
    cats = Reference(ws, min_col=1, min_row=12, max_row=15)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 11
    ws.add_chart(chart, "A18")

    chart2 = BarChart()
    chart2.title = "Prioridade do lote"
    chart2.style = 12
    chart2.y_axis.title = "Quantidade"
    data2 = Reference(ws, min_col=5, min_row=10, max_row=13)
    cats2 = Reference(ws, min_col=4, min_row=11, max_row=13)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.height = 7
    chart2.width = 8
    ws.add_chart(chart2, "M18")

    ws["M5"] = "Uso recomendado"
    style_cell(ws["M5"], fill=header_fill, font=Font(color="FFFFFF", bold=True))
    notes = [
        "1. Revisar o outreach pack antes de qualquer envio.",
        "2. Priorizar os casos com e-mail corporativo claro.",
        "3. Usar WhatsApp inicialmente em modo manual assistido.",
        "4. Atualizar status outbound e datas a cada toque.",
    ]
    for i, note in enumerate(notes, start=6):
        ws[f"M{i}"] = note
        ws[f"M{i}"].alignment = Alignment(wrap_text=True)

    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 18
    for col in ["M", "N", "O", "P"]:
        ws.column_dimensions[col].width = 22


def build_leads(ws, rows: list[dict[str, str]]):
    ws.freeze_panes = "A2"
    headers = [
        "company_id",
        "company_name",
        "segment",
        "website_url",
        "instagram_url",
        "public_email",
        "public_phone",
        "public_whatsapp",
        "normalized_whatsapp",
        "whatsapp_readiness",
        "contact_type",
        "business_summary",
        "digital_presence_stage",
        "suggested_offer_primary",
        "suggested_offer_secondary",
        "priority",
        "recommended_channel",
        "fallback_contact_channel",
        "contact_readiness",
        "confidence_business",
        "confidence_contact",
        "confidence_offer_fit",
        "review_status",
        "outbound_status",
        "owner_human",
        "first_touch_date",
        "last_touch_date",
        "next_follow_up_date",
        "reply_status",
        "notes",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor=BRAND)
    border = Border(
        left=Side(style="thin", color=LINE),
        right=Side(style="thin", color=LINE),
        top=Side(style="thin", color=LINE),
        bottom=Side(style="thin", color=LINE),
    )
    for cell in ws[1]:
        style_cell(
            cell,
            fill=header_fill,
            font=Font(color="FFFFFF", bold=True),
            align=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=border,
        )

    for row in rows:
        ws.append(
            [
                row["company_id"],
                row["company_name"],
                row["segment"],
                row["website_url"],
                row["instagram_url"],
                row["public_email"],
                row["public_phone"],
                row["public_whatsapp"],
                row.get("normalized_whatsapp", ""),
                row.get("whatsapp_readiness", ""),
                row["contact_type"],
                row["business_summary"],
                row["digital_presence_stage"],
                row["suggested_offer_primary"],
                row["suggested_offer_secondary"],
                row["priority"],
                row["recommended_channel"],
                row.get("fallback_contact_channel", ""),
                row.get("contact_readiness", ""),
                int(row["confidence_business"]),
                int(row["confidence_contact"]),
                int(row["confidence_offer_fit"]),
                row["review_status"],
                row["outbound_status"],
                row["owner_human"],
                row["first_touch_date"],
                row["last_touch_date"],
                row["next_follow_up_date"],
                row["reply_status"],
                row["notes"],
            ]
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    add_table(ws, "A1", f"AD{len(rows) + 1}", "LeadsTable")

    dv_status = DataValidation(
        type="list",
        formula1='"nao-iniciado,enviado,aguardando-resposta,respondeu,sem-resposta,nao-contatar"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_status)
    dv_status.add(f"X2:X{len(rows) + 20}")

    dv_review = DataValidation(
        type="list",
        formula1='"novo,revisando,pronta-para-contato,bloqueado,contatada"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_review)
    dv_review.add(f"W2:W{len(rows) + 20}")

    ws.conditional_formatting.add(
        f"P2:P{len(rows) + 20}",
        CellIsRule(operator="equal", formula=['"alta"'], fill=PatternFill("solid", fgColor=YELLOW)),
    )
    ws.conditional_formatting.add(
        f"X2:X{len(rows) + 20}",
        CellIsRule(operator="equal", formula=['"respondeu"'], fill=PatternFill("solid", fgColor=GREEN)),
    )
    ws.conditional_formatting.add(
        f"X2:X{len(rows) + 20}",
        CellIsRule(operator="equal", formula=['"sem-resposta"'], fill=PatternFill("solid", fgColor=RED)),
    )
    ws.conditional_formatting.add(
        f"X2:X{len(rows) + 20}",
        CellIsRule(operator="equal", formula=['"enviado"'], fill=PatternFill("solid", fgColor=BLUE)),
    )

    for cell in ws["D"][1:]:
        if cell.value:
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"
    for cell in ws["E"][1:]:
        if cell.value:
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"

    widths = {
        "A": 12,
        "B": 28,
        "C": 22,
        "D": 28,
        "E": 26,
        "F": 24,
        "G": 16,
        "H": 18,
        "I": 18,
        "J": 18,
        "K": 14,
        "L": 18,
        "M": 36,
        "N": 16,
        "O": 22,
        "P": 22,
        "Q": 12,
        "R": 22,
        "S": 18,
        "T": 16,
        "U": 12,
        "V": 12,
        "W": 18,
        "X": 18,
        "Y": 16,
        "Z": 16,
        "AA": 16,
        "AB": 18,
        "AC": 16,
        "AD": 28,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def build_outreach(ws, rows: list[dict[str, str]]):
    ws.freeze_panes = "A2"
    headers = [
        "company_id",
        "company_name",
        "recommended_channel",
        "email_subject",
        "email_body",
        "whatsapp_body",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor=BRAND)
    border = Border(
        left=Side(style="thin", color=LINE),
        right=Side(style="thin", color=LINE),
        top=Side(style="thin", color=LINE),
        bottom=Side(style="thin", color=LINE),
    )
    for cell in ws[1]:
        style_cell(
            cell,
            fill=header_fill,
            font=Font(color="FFFFFF", bold=True),
            align=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=border,
        )

    for row in rows:
        ws.append(
            [
                row["company_id"],
                row["company_name"],
                row["recommended_channel"],
                row["email_subject"],
                row["email_body"],
                row["whatsapp_body"],
            ]
        )
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    add_table(ws, "A1", f"F{len(rows) + 1}", "OutreachTable")
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 70
    ws.column_dimensions["F"].width = 70
    for row_idx in range(2, len(rows) + 2):
        ws.row_dimensions[row_idx].height = 95


def build_cadence(ws):
    ws["A1"] = "Cadencia sugerida"
    ws["A2"] = "Dia"
    ws["B2"] = "Acao"
    ws["C2"] = "Observacao"
    header_fill = PatternFill("solid", fgColor=BRAND)
    border = Border(
        left=Side(style="thin", color=LINE),
        right=Side(style="thin", color=LINE),
        top=Side(style="thin", color=LINE),
        bottom=Side(style="thin", color=LINE),
    )
    style_cell(
        ws["A1"],
        font=Font(bold=True, size=14, color=INK),
    )
    for cell in ws[2]:
        style_cell(
            cell,
            fill=header_fill,
            font=Font(color="FFFFFF", bold=True),
            align=Alignment(horizontal="center"),
            border=border,
        )
    steps = [
        ("D0", "Primeiro toque", "Enviar e-mail ou WhatsApp conforme canal recomendado."),
        ("D3", "Follow-up 1", "Retomar de forma curta, sem repetir o texto inteiro."),
        ("D7", "Follow-up 2", "Trazer novo angulo ou exemplo pratico da oferta."),
        ("D14", "Encerramento", "Encerrar com abertura para contato futuro e marcar sem resposta."),
    ]
    for row in steps:
        ws.append(row)
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 60


def main():
    rows = load_rows()
    wb = Workbook()
    wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
    build_dashboard(wb, len(rows))
    leads = wb.create_sheet("Leads")
    outreach = wb.create_sheet("Outreach")
    cadence = wb.create_sheet("Cadencia")
    build_leads(leads, rows)
    build_outreach(outreach, rows)
    build_cadence(cadence)
    OUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_FILE)
    print(OUT_FILE)


if __name__ == "__main__":
    main()
